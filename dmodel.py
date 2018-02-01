import csv, json, yaml
import os, sys

from subprocess import run
run("pip3 install openpyxl", shell=True, check=True)
from openpyxl import load_workbook

# /TODO Create cli args to convert *.X files in an entire directory to *.Y files
# /TODO Create feature that converts any any instead of just JSON output

def main(file):

    # Separates EXCEL worksheets from a woorkbook
    def XLSXtoCSV(object):
        def get_all_sheets(object):
            sheets = []
            workbook = load_workbook(object, read_only=True, data_only=True)
            all_worksheets = workbook.sheetnames
            for worksheet_name in all_worksheets:
                sheets.append(worksheet_name)
            return sheets

        def csv_from_excel(object, sheets):
            workbook = load_workbook(object, data_only=True)
            for worksheet_name in sheets:
                print("Export " + worksheet_name + " ...")
                try:
                    worksheet = workbook[(worksheet_name)]
                except KeyError:
                    print("Could not find " + worksheet_name)
                    sys.exit(1)

                your_csv_file = open(''.join([worksheet_name, '.csv']), newline="", mode='wt')
                wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)
                for row in worksheet.iter_rows():
                    lrow = []
                    for cell in row:
                        lrow.append(cell.value)
                    wr.writerow(lrow)

                print(" ... done")
                your_csv_file.close()
        if not 2 <= len(sys.argv) <= 3:
            print("Call with " + sys.argv[0] + " <xlxs file> [comma separated list of sheets to export]")
            sys.exit(1)
        else:
            sheets = []
            if len(sys.argv) == 3:
                sheets = list(sys.argv[2].split(','))
            else:
                sheets = get_all_sheets(sys.argv[1])
            assert (sheets != None and len(sheets) > 0)
            csv_from_excel(sys.argv[1], sheets)

    # Convert file to JSON
    def CSVtoJSON(object):
        for CSV in object:
            CSVfile = os.path.splitext(object)[0]
            JSONfile = CSVfile + '.JSON'

            with open(CSVfile + '.CSV') as f:
                reader = csv.DictReader(f)
                rows = list(reader)

            with open(JSONfile, 'w', ) as f:
                json.dump(rows, f, indent=4)

    # Converts YAML files to JSON
    def YAMLtoJSON(object):
        pass

    # Converts JSON files to ???
    def JSONtoYAML(object):
        #YAMLfile = open(os.path.splitext(object)[0] + '.yml', 'w')
        #yaml.dump(json.loads((object)), YAMLfile, default_flow_style=False)

        # "json.decoder.JSONDecodeError: Expecting value: line 1 column 1 (char 0)"
        pass

    # Regex to determine file type, returns tuple with two values (the filename without extension + just the extension).
    # The second index ([1]) gives us just the extension. The cool thing is, that this way you can also access the
    # filename pretty easily, if needed!

    if os.path.splitext(file)[1] == ".xlsx":
        print('Excel File.Converting each worksheet in workbook to JSON.')
        XLSXtoCSV(file)
    elif os.path.splitext(file)[1] == ".csv":
        print('CSV file. Converting to JSON.')
        CSVtoJSON(file)
    elif os.path.splitext(file)[1].lower() == ".json":
        #print('JSON file. Converting to YAML.')
        #JSONtoYAML(file)
        print('JSON file. Feature not yet implemented, sorry...')
        pass
    elif os.path.splitext(file)[1] == ".yaml" or '.yml':
        #print('YAML file. Converting to JSON.')
        #YAMLtoJSON(file)
        print('YAML file. Feature not yet implemented, sorry...')
        pass
    else:
        print('Unsupported file type. Try: .xlsx, .csv, .yaml, .yml or .json')
        sys.exit(1)

if __name__ == '__main__':
    file = sys.argv[1] if len(sys.argv) > 1 else print('Usage: "python dmodel.py data.csv" ') / sys.exit(0)
    main(file)